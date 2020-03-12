from openpyxl import load_workbook
excel_document = load_workbook('VMWare&OracleData.xlsx')

print(type(excel_document))
sheet = excel_document.sheetnames
range_names = excel_document['VMWare&OracleData']
print(sheet[0])
sheetRange = 'A2:A4623'
ClientLinuxInstances = {}
clientCount = {}
for r in range_names[sheetRange]:
    if r[0].value != 'NULL':
        queryValue = str(r[0].value)
        if str(r[0].value) not in clientCount:
            clientCount[r[0].value] = 1
        if str(r[0].value) not in ClientLinuxInstances:
            ClientLinuxInstances[r[0].value] = {'OS' : range_names['s' + str(r[0].row)].value, 'Count' : clientCount[r[0].value] }
        else :
            clientCount[r[0].value] = clientCount[r[0].value] + 1
            ClientLinuxInstances[r[0].value] = {'OS' : range_names['s' + str(r[0].row)].value, 'Count' : clientCount[r[0].value] }
        print(r[0].value)
        print(range_names['s' + str(r[0].row)].value)

LOG('Complete')